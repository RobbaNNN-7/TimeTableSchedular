from typing import List, Dict, Tuple
import random
from openpyxl import Workbook # type: ignore
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font # type: ignore # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

class TimeTableCSP:
    def __init__(self):
        self.sections = ['CSE-A', 'CSE-B', 'CSE-C', 'CSE-D', 'CSE-E']
        self.days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        self.hours_per_day = 8
        self.break_hour = 4  # 5th hour (0-based index is 4)
        
        # Courses with format: {name: {'theory': hours, 'lab': hours}}
        self.courses = {
            'Data Structures': {'theory': 3, 'lab': 0},
            'Programming': {'theory': 3, 'lab': 3},
            'Algorithms': {'theory': 3, 'lab': 0},
            'Database': {'theory': 3, 'lab': 0},
            'Networks': {'theory': 2, 'lab': 0},
            'Operating Systems': {'theory': 2, 'lab':3 }
        }
        
        self.theory_rooms = ['Room101', 'Room102', 'Room103', 'Room104', 'Room105']
        self.lab_rooms = ['Lab101', 'Lab102', 'Lab103', 'Lab104', 'Lab105']
        
        # Initialize empty timetable
        self.timetable = {
            section: {
                day: [None for _ in range(self.hours_per_day)]
                for day in range(len(self.days))
            }
            for section in self.sections
        }

        self.room_schedule = {

            room:{   
                day: [False] * self.hours_per_day
                for day in range(len(self.days))
            }

            for room in (self.theory_rooms + self.lab_rooms)
        }
        
        # Set break time for all sections
        self._set_break_time()

    def _set_break_time(self):
        """Set break time for all sections on all days"""
        for section in self.sections:
            for day in range(len(self.days)):
                self.timetable[section][day][self.break_hour] = {
                    'course': 'Break',
                    'room': 'Break',
                    'type': 'break'
                }

    def is_slot_free(self, section: str, day: int, hour: int, duration: int = 1,room : str = None) -> bool:
        """Check if a time slot is free for the given duration"""
        if hour + duration > self.hours_per_day:
            return False
            
        # Check if the slot overlaps with break time
        if any(h == self.break_hour for h in range(hour, hour + duration)):
            return False
        
        # Checking Slot Availibility
        if not all(self.timetable[section][day][h] is None 
                  for h in range(hour, hour + duration)):
                    return False
        
        # Checking Room Availibility
        if room:
            if not all(not self.room_schedule[room][day][h]
            for h in range(hour,hour+duration)):
                return False

        """ Availible """
        return True



    def find_free_slot_and_room(self, section: str, duration: int = 1, is_lab: bool = False) -> Tuple[int, int, str]:
        """Find a free slot and room with intelligent backtracking"""
        # Randomize search order to increase solution probability
        days = list(range(len(self.days)))
        hours = list(range(self.hours_per_day - duration + 1))
        rooms = self.lab_rooms if is_lab else self.theory_rooms

        random.shuffle(days)
        random.shuffle(hours)
        random.shuffle(rooms)

        for day in days:
            for hour in hours:
                for room in rooms:
                    if self.is_slot_free(section, day, hour, duration, room):
                        return day, hour, room
        
        return None

    def schedule_session(self, section: str, course: str, is_lab: bool)->bool:
        """Schedule a theory or lab session"""
        duration = 3 if is_lab else 1
        course_name = f"{course} Lab" if is_lab else course
        
        slot = self.find_free_slot_and_room(section, duration,is_lab)
        if not slot:
            return False # Not Possible To Assign
        
        day, start_hour ,room = slot
        
        # Schedule the session
        for hour in range(start_hour, start_hour + duration):
            self.timetable[section][day][hour] = {
                'course': course_name,
                'room': room,
                'type': 'lab' if is_lab else 'theory'
            }
            self.room_schedule[room][day][hour] = True # Room Scheduled
        return True

    def generate_timetable(self) -> bool:
        """Improved timetable generation with backtracking"""
        # First schedule labs (require 3 consecutive hours)
        for course, hours in self.courses.items():
            if hours['lab'] > 0:
                lab_scheduled = all(
                    self.schedule_session(section, course, is_lab=True)
                    for section in self.sections
                )
                if not lab_scheduled:
                    return False
        
        # Then schedule theory classes
        for course, hours in self.courses.items():
            for _ in range(hours['theory']):
                theory_scheduled = all(
                    self.schedule_session(section, course, is_lab=False)
                    for section in self.sections
                )
                if not theory_scheduled:
                    return False
        
        return True

        

    def solve(self):
        max_attempts = 10000
        for attempt in range(max_attempts):
            # Reset timetable
            self.timetable = {
                section: {
                    day: [None for _ in range(self.hours_per_day)]
                    for day in range(len(self.days))
                }
                for section in self.sections
            }

            # reset rooms

            
            self.room_schedule = {
                room: {   
                    day: [False] * self.hours_per_day
                    for day in range(len(self.days))
                }
                for room in (self.theory_rooms + self.lab_rooms)
            }
            
            # Set break time before generating schedule
            self._set_break_time()
            
            if self.generate_timetable():
                return True
        return False

    def export_to_excel(self, filename: str = 'timetable.xlsx'):
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        theory_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        lab_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        break_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(bold=True)
        
        for section in self.sections:
            if section == self.sections[0]:
                ws = wb.active
                ws.title = section
            else:
                ws = wb.create_sheet(section)
            
            # Set column widths
            ws.column_dimensions['A'].width = 10
            for i in range(len(self.days)):
                col = get_column_letter(i + 2)
                ws.column_dimensions[col].width = 20
            
            # Write section name
            ws.merge_cells('A1:F1')
            cell = ws['A1']
            cell.value = f"Timetable for {section}"
            cell.fill = header_fill
            cell.font = Font(bold=True, size=14)
            cell.alignment = center_aligned
            cell.border = border
            
            # Write days
            for day_idx, day in enumerate(self.days):
                cell = ws.cell(row=2, column=day_idx + 2)
                cell.value = day
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_aligned
            
            # Write hours
            for hour in range(self.hours_per_day):
                cell = ws.cell(row=hour + 3, column=1)
                cell.value = f"Hour {hour + 1}"
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_aligned
            
            # Fill in the schedule
            for day in range(len(self.days)):
                for hour in range(self.hours_per_day):
                    cell = ws.cell(row=hour + 3, column=day + 2)
                    slot = self.timetable[section][day][hour]
                    
                    if slot:
                        cell.value = f"{slot['course']}\n{slot['room']}"
                        if slot['type'] == 'break':
                            cell.fill = break_fill
                        elif slot['type'] == 'lab':
                            cell.fill = lab_fill
                        else:
                            cell.fill = theory_fill
                    else:
                        cell.value = "---"
                    
                    cell.border = border
                    cell.alignment = center_aligned
            
            # Set row heights
            ws.row_dimensions[1].height = 30
            for row in range(3, 11):
                ws.row_dimensions[row].height = 60
        
        wb.save(filename)
        print(f"Timetable has been exported to {filename}")

scheduler = TimeTableCSP()
if scheduler.solve():
    scheduler.export_to_excel("optimized_timetable.xlsx")
    print("Timetable generated successfully!")
else:
    print("Could not generate a valid timetable after maximum attempts.")