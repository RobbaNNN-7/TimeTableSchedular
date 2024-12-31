# timetable_generator.py
from typing import List, Dict, Tuple, Optional
import random
from openpyxl import Workbook # type: ignore
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
import json

class TimeTableCSP:
    def __init__(self, sections=None, courses=None, theory_rooms=None, lab_rooms=None):
        self.sections = sections or ['CSE-A', 'CSE-B', 'CSE-C', 'CSE-D', 'CSE-E']
        self.days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        self.hours_per_day = 8
        self.break_hour = 4
        
        self.courses = courses or {
            'Data Structures': {'theory': 3, 'lab': 0},
            'Programming': {'theory': 3, 'lab': 3},
            'Algorithms': {'theory': 3, 'lab': 0},
            'Database': {'theory': 3, 'lab': 0},
            'Networks': {'theory': 2, 'lab': 0},
            'Operating Systems': {'theory': 2, 'lab': 3}
        }
        
        self.theory_rooms = theory_rooms or ['Room101', 'Room102', 'Room103', 'Room104', 'Room105']
        self.lab_rooms = lab_rooms or ['Lab101', 'Lab102', 'Lab103', 'Lab104', 'Lab105']
        
        self.initialize_timetable()
        self.monte_carlo_stats = {
            'total_attempts': 0,
            'successful_attempts': 0,
            'failed_attempts': 0
        }

    def initialize_timetable(self):
        self.timetable = {
            section: {
                day: [None for _ in range(self.hours_per_day)]
                for day in range(len(self.days))
            }
            for section in self.sections
        }
        
        self.room_schedule = {
            room: {
                day: [False] * self.hours_per_day
                for day in range(len(self.days))
            }
            for room in (self.theory_rooms + self.lab_rooms)
        }
        self._set_break_time()

    def _set_break_time(self):
        for section in self.sections:
            for day in range(len(self.days)):
                self.timetable[section][day][self.break_hour] = {
                    'course': 'Break',
                    'room': 'Break',
                    'type': 'break'
                }

    def is_consecutive_slots(self, section: str, day: int, hour: int, course: str) -> bool:
        for start in range(hour - 2, hour + 1):
            if 0 <= start <= self.hours_per_day - 3:
                slots = [
                    self.timetable[section][day][h]
                    for h in range(start, start + 3)
                    if h != self.break_hour
                ]
                if all(slot and slot['course'] == course for slot in slots):
                    return False
        return True

    def check_gaps(self, section: str, day: int, hour: int, course: str) -> bool:
        schedule = self.timetable[section][day]
        course_slots = [
            i for i, slot in enumerate(schedule)
            if slot and slot['course'] == course and i != hour
        ]
        
        if not course_slots:
            return True
            
        min_slot = min(course_slots)
        max_slot = max(course_slots)
        
        if hour < min_slot:
            return (min_slot - hour) <= 2
        if hour > max_slot:
            return (hour - max_slot) <= 2
            
        return True

    def is_slot_free(self, section: str, day: int, hour: int, course: str = None, duration: int = 1) -> bool:
        if hour + duration > self.hours_per_day:
            return False
            
        if any(h == self.break_hour for h in range(hour, hour + duration)):
            return False
        
        if not all(self.timetable[section][day][h] is None 
                  for h in range(hour, hour + duration)):
            return False
        
        if course and not self.is_consecutive_slots(section, day, hour, course):
            return False
            
        if course and not self.check_gaps(section, day, hour, course):
            return False
            
        return True

    def find_free_slot_and_room(self, section: str, course: str, is_lab: bool = False) -> Optional[Tuple[int, int, str]]:
        duration = 3 if is_lab else 1
        days = list(range(len(self.days)))
        hours = list(range(self.hours_per_day - duration + 1))
        rooms = self.lab_rooms if is_lab else self.theory_rooms
        
        random.shuffle(days)
        random.shuffle(hours)
        random.shuffle(rooms)
        
        for day in days:
            for hour in hours:
                if self.is_slot_free(section, day, hour, course, duration):
                    for room in rooms:
                        if all(not self.room_schedule[room][day][h] 
                              for h in range(hour, hour + duration)):
                            return day, hour, room
        return None

    def schedule_session(self, section: str, course: str, is_lab: bool) -> bool:
        slot = self.find_free_slot_and_room(section, course, is_lab)
        if not slot:
            return False
            
        day, start_hour, room = slot
        duration = 3 if is_lab else 1
        course_name = f"{course} Lab" if is_lab else course
        
        for hour in range(start_hour, start_hour + duration):
            self.timetable[section][day][hour] = {
                'course': course_name,
                'room': room,
                'type': 'lab' if is_lab else 'theory'
            }
            self.room_schedule[room][day][hour] = True
        return True

    def generate_timetable(self) -> bool:
        # Schedule labs first
        for course, hours in self.courses.items():
            if hours['lab'] > 0:
                lab_scheduled = all(
                    self.schedule_session(section, course, is_lab=True)
                    for section in self.sections
                )
                if not lab_scheduled:
                    return False
        
        # Then schedule theory classes
        for course, hours in self.courses.items():
            for _ in range(hours['theory']):
                theory_scheduled = all(
                    self.schedule_session(section, course, is_lab=False)
                    for section in self.sections
                )
                if not theory_scheduled:
                    return False
        
        return True

    def add_makeup_class(self, section: str, course: str, is_lab: bool = False) -> bool:
        slot = self.find_free_slot_and_room(section, course, is_lab)
        if not slot:
            return False
            
        day, hour, room = slot
        duration = 3 if is_lab else 1
        course_name = f"{course} Makeup {'Lab' if is_lab else 'Class'}"
        
        for h in range(hour, hour + duration):
            self.timetable[section][day][h] = {
                'course': course_name,
                'room': room,
                'type': 'makeup'
            }
            self.room_schedule[room][day][h] = True
        return True

    def monte_carlo_simulation(self, num_iterations: int = 1000) -> Dict:
        best_score = float('-inf')
        best_timetable = None
        best_room_schedule = None
        
        for _ in range(num_iterations):
            self.monte_carlo_stats['total_attempts'] += 1
            self.initialize_timetable()
            
            if self.generate_timetable():
                self.monte_carlo_stats['successful_attempts'] += 1
                current_score = self.evaluate_timetable()
                if current_score > best_score:
                    best_score = current_score
                    best_timetable = self.copy_timetable()
                    best_room_schedule = self.copy_room_schedule()
            else:
                self.monte_carlo_stats['failed_attempts'] += 1

        if best_timetable:
            self.timetable = best_timetable
            self.room_schedule = best_room_schedule
            return {
                'success_rate': (self.monte_carlo_stats['successful_attempts'] / 
                               self.monte_carlo_stats['total_attempts']) * 100,
                'best_score': best_score
            }
        return {'success_rate': 0, 'best_score': 0}

    def evaluate_timetable(self) -> float:
        score = 0
        for section in self.sections:
            for day in range(len(self.days)):
                gaps = self.count_gaps(section, day)
                score -= gaps * 10
                distribution = self.evaluate_distribution(section, day)
                score += distribution
        return score

    def count_gaps(self, section: str, day: int) -> int:
        schedule = self.timetable[section][day]
        class_slots = [
            i for i, slot in enumerate(schedule)
            if slot and slot['type'] not in ['break', 'makeup']
        ]
        
        if len(class_slots) <= 1:
            return 0
            
        gaps = 0
        for i in range(len(class_slots) - 1):
            gap = class_slots[i + 1] - class_slots[i] - 1
            if gap > 1:  # Allow one hour gaps
                gaps += gap - 1
        return gaps

    def evaluate_distribution(self, section: str, day: int) -> float:
        schedule = self.timetable[section][day]
        class_slots = [
            i for i, slot in enumerate(schedule)
            if slot and slot['type'] not in ['break', 'makeup']
        ]
        
        if len(class_slots) <= 1:
            return 0
            
        gaps = [class_slots[i + 1] - class_slots[i]
                for i in range(len(class_slots) - 1)]
        avg_gap = sum(gaps) / len(gaps)
        variance = sum((gap - avg_gap) ** 2 for gap in gaps) / len(gaps)
        
        return 10 / (1 + variance)

    def copy_timetable(self):
        return {
            section: {
                day: [slot.copy() if slot else None 
                      for slot in day_schedule]
                for day, day_schedule in section_schedule.items()
            }
            for section, section_schedule in self.timetable.items()
        }

    def copy_room_schedule(self):
        return {
            room: {
                day: list(schedule)
                for day, schedule in days.items()
            }
            for room, days in self.room_schedule.items()
        }

    def export_to_excel(self, filename: str = 'timetable.xlsx'):
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        theory_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        lab_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        break_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        makeup_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(bold=True)
        
        for section in self.sections:
            if section == self.sections[0]:
                ws = wb.active
                ws.title = section
            else:
                ws = wb.create_sheet(section)
            
            # Set column widths
            ws.column_dimensions['A'].width = 10
            for i in range(len(self.days)):
                col = get_column_letter(i + 2)
                ws.column_dimensions[col].width = 20
            
            # Write headers
            ws.merge_cells(f'A1:{get_column_letter(len(self.days) + 1)}1')
            cell = ws['A1']
            cell.value = f"Timetable for {section}"
            cell.fill = header_fill
            cell.font = Font(bold=True, size=14)
            cell.alignment = center_aligned
            
            # Write days
            for day_idx, day in enumerate(self.days):
                cell = ws.cell(row=2, column=day_idx + 2)
                cell.value = day
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_aligned
            
            # Write hours
            for hour in range(self.hours_per_day):
                cell = ws.cell(row=hour + 3, column=1)
                cell.value = f"Hour {hour + 1}"
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_aligned
            
            # Fill schedule
            for day in range(len(self.days)):
                for hour in range(self.hours_per_day):
                    cell = ws.cell(row=hour + 3, column=day + 2)
                    slot = self.timetable[section][day][hour]
                    
                    if slot:
                        cell.value = f"{slot['course']}\n{slot['room']}"
                        if slot['type'] == 'break':
                            cell.fill = break_fill
                        elif slot['type'] == 'lab':
                            cell.fill = lab_fill
                        elif slot['type'] == 'makeup':
                            cell.fill = makeup_fill
                        else:
                            cell.fill = theory_fill
                    else:
                        cell.value = "---"
                    
                    cell.border = border
                    cell.alignment = center_aligned
            
            # Set row heights
            for row in range(1, self.hours_per_day + 3):
                ws.row_dimensions[row].height = 40
        
        wb.save(filename)
        return f"Timetable exported to {filename}"

if __name__ == "__main__":
    scheduler = TimeTableCSP()
    results = scheduler.monte_carlo_simulation()
    if results['success_rate'] > 0:
        scheduler.export_to_excel()
        print(f"Timetable generated successfully!")
        print(f"Success rate: {results['success_rate']:.2f}%")
        print(f"Best score: {results['best_score']:.2f}")
    else:
        print("Failed to generate timetable")